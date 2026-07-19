#!/usr/bin/env python3
"""Rebuild the Teacher Time Table from a (manually edited) Class Time Table.

Parses the Class sheet verbatim (honouring your edits — no re-solving), regenerates
the Teacher sheet to match, preserves the Class sheet exactly, checks for teacher
double-booking, and writes both the .xlsx and a .pdf.

Usage:
  python rebuild_teacher_tt.py --in <file.xlsx> --school NRCS \
      --out-xlsx <out.xlsx> --out-pdf <out.pdf>
"""
import argparse
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment

from timetable.model import DAYS, N_DAYS, STUDY_PERIOD, SCHOOLS, Model
from timetable.writer import (NAVY, DAYFILL, PERFILL, STUDYFILL, TEACHHDR,
                              WHITE_BOLD, _cell)
from timetable.pdf import write_pdf

PARALLEL = {"P.E", "MARTIAL ARTS", "Martial Arts", "P.E.T Instructor", "Karate Instructor"}


def _append(day_row, period, text):
    """Accumulate multiple classes in one slot (combined P.E/Karate sessions)."""
    day_row[period] = f"{day_row[period]}\n{text}" if day_row[period] else text


def parse_class_sheet(ws):
    """-> (tgrid, solution, supervisors, classes, teachers)."""
    classes = {c: str(ws.cell(1, c).value).strip()
               for c in range(3, ws.max_column + 1) if ws.cell(1, c).value}
    tgrid = defaultdict(lambda: [["" for _ in range(9)] for _ in range(N_DAYS)])
    solution, supervisors, teachers = {}, {}, set()
    day = -1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            day += 1
        plabel = ws.cell(r, 2).value
        if plabel in (None, ""):
            continue
        is_study = str(plabel).strip().lower().startswith("study")
        period = STUDY_PERIOD if is_study else int(plabel)
        for c, cls in classes.items():
            raw = ws.cell(r, c).value
            if not raw:
                continue
            s = str(raw).strip()
            if s.lower() in ("recreation", "free period", "—", "-"):
                continue
            parts = s.split("\n")
            teacher = parts[0].strip()
            subj = parts[1].strip().strip("()").strip() if len(parts) > 1 else ""
            teachers.add(teacher)
            if subj.upper() == "CLASS" or (is_study and not subj):
                supervisors[cls] = teacher
                _append(tgrid[teacher][day], STUDY_PERIOD, cls)
            else:
                _append(tgrid[teacher][day], period, f"{cls} ({subj})" if subj else cls)
                solution[(cls, day, period)] = (subj, teacher)
    return tgrid, solution, supervisors, list(classes.values()), teachers


def check_conflicts(solution, supervisors):
    occ = defaultdict(list)
    for (c, d, p), (s, t) in solution.items():
        if t in PARALLEL:
            continue
        occ[(t, d, p)].append(f"{c}({s})")
    for c, t in supervisors.items():          # study-hour supervision occupies P8
        for d in range(N_DAYS):
            occ[(t, d, STUDY_PERIOD)].append(f"{c}(STUDY)")
    return [f"{t} @ {DAYS[d]} P{p}: {v}"
            for (t, d, p), v in sorted(occ.items()) if len(v) > 1]


def write_teacher_sheet(wb, cfg, tgrid, teachers):
    if cfg.sheet_teacher_name in wb.sheetnames:
        del wb[cfg.sheet_teacher_name]
    ws = wb.create_sheet(cfg.sheet_teacher_name, 0)
    _cell(ws, 1, 1, cfg.teacher_title, NAVY, Font(bold=True, color="FFFFFF", size=13))
    ws.merge_cells("A1:I1")
    real = sorted(t for t in teachers if t not in PARALLEL)
    generic = sorted(t for t in teachers if t in PARALLEL)
    row = 3
    for t in real + generic:
        _cell(ws, row, 1, f"TEACHER: {t}", TEACHHDR, Font(bold=True, color="FFFFFF", size=11),
              Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        for j, h in enumerate(["DAY", "PERIOD 1", "PERIOD 2", "PERIOD 3", "PERIOD 4",
                               "PERIOD 5", "PERIOD 6", "PERIOD 7", "STUDY HOUR"], start=1):
            _cell(ws, row, j, h, DAYFILL, WHITE_BOLD)
        row += 1
        for d in range(N_DAYS):
            _cell(ws, row, 1, DAYS[d], PERFILL, Font(bold=True, size=10))
            for p in range(1, 9):
                txt = tgrid[t][d][p]
                col = 9 if p == STUDY_PERIOD else p + 1
                _cell(ws, row, col, txt, STUDYFILL if (p == STUDY_PERIOD and txt) else None)
            row += 1
        row += 1
    ws.column_dimensions["A"].width = 10
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["I"].width = 15


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--school", default="NRCS")
    ap.add_argument("--out-xlsx", required=True)
    ap.add_argument("--out-pdf", required=True)
    args = ap.parse_args()
    cfg = SCHOOLS[args.school]

    wb = openpyxl.load_workbook(args.inp)          # preserve Class sheet + styling
    cs = wb[cfg.sheet_class_name]
    tgrid, solution, supervisors, classes, teachers = parse_class_sheet(cs)

    conflicts = check_conflicts(solution, supervisors)
    print(f"Parsed: {len(classes)} classes, {len(teachers)} teachers, {len(solution)} taught slots.")
    print(f"Study-hour supervisors: {len(supervisors)} classes.")
    if conflicts:
        print(f"\n⚠ {len(conflicts)} DOUBLE-BOOKING(S) in the Class sheet:")
        for c in conflicts:
            print("   ", c)
    else:
        print("✅ No teacher double-booking in the Class sheet.")

    write_teacher_sheet(wb, cfg, tgrid, teachers)
    wb.save(args.out_xlsx)
    print(f"\nWrote {args.out_xlsx}")

    m = Model(cfg=cfg, classes=classes, subjects=[], plan={}, teacher_of={},
              p1_teacher={}, study_supervisor=supervisors,
              teachers=sorted(t for t in teachers if t not in PARALLEL),
              study_hour_classes=list(supervisors), subjects_of={})
    write_pdf(args.out_pdf, m, solution)
    print(f"Wrote {args.out_pdf}")
    return 1 if conflicts else 0


if __name__ == "__main__":
    raise SystemExit(main())
