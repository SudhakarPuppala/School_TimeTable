"""Write the styled workbook: Class Time Table + Teacher Time Table (per m.cfg)."""
from __future__ import annotations
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .model import Model, DAYS, N_DAYS, STUDY_PERIOD

NAVY = PatternFill("solid", fgColor="2A4D69")
DAYFILL = PatternFill("solid", fgColor="4B86B4")
PERFILL = PatternFill("solid", fgColor="E7EFF6")
STUDYFILL = PatternFill("solid", fgColor="FCE9D6")
FREEFILL = PatternFill("solid", fgColor="F2F2F2")
TEACHHDR = PatternFill("solid", fgColor="000080")
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)
FREE_FONT = Font(size=10, italic=True, color="808080")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_ROT = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
DAY_FULL = {"MON": "MONDAY", "TUE": "TUESDAY", "WED": "WEDNESDAY",
            "THU": "THURSDAY", "FRI": "FRIDAY", "SAT": "SATURDAY"}
ROW_LABELS = [1, 2, 3, 4, 5, 6, 7, "Study Hour"]


def _abbr(cfg, subj):
    return cfg.subj_abbr.get(subj, subj[:4].upper())


def _cell(ws, r, c, value, fill=None, font=None, align=CENTER):
    cell = ws.cell(r, c, value)
    if fill:
        cell.fill = fill
    cell.font = font or CELL_FONT
    cell.alignment = align
    cell.border = BORDER
    return cell


def write_class_sheet(wb, m: Model, solution):
    cfg = m.cfg
    ws = wb.create_sheet(cfg.sheet_class_name)
    _cell(ws, 1, 1, "Day", NAVY, WHITE_BOLD)
    _cell(ws, 1, 2, "Period", NAVY, WHITE_BOLD)
    for j, c in enumerate(m.classes, start=3):
        _cell(ws, 1, j, cfg.class_display.get(c, c), NAVY, WHITE_BOLD)

    row = 2
    for d in range(N_DAYS):
        day_start = row
        for label in ROW_LABELS:
            p = STUDY_PERIOD if label == "Study Hour" else label
            is_study = label == "Study Hour"
            _cell(ws, row, 2, label, PERFILL, Font(bold=True, size=10))
            for j, c in enumerate(m.classes, start=3):
                fill, font = None, CELL_FONT
                if is_study and not m.has_p8(DAYS[d]):
                    text, fill = "", FREEFILL              # no Period 8 this day
                elif is_study and c in m.study_hour_classes:
                    ct = m.study_supervisor.get(c, "")
                    text, fill = f"{ct}\n(CLASS)", STUDYFILL
                elif (c, d, p) in solution:
                    s, t = solution[(c, d, p)]
                    text = f"{t}\n({_abbr(cfg, s)})"
                    if is_study:
                        fill = STUDYFILL
                else:
                    text, fill, font = "Recreation", FREEFILL, FREE_FONT
                _cell(ws, row, j, text, fill, font)
            row += 1
        ws.merge_cells(start_row=day_start, start_column=1, end_row=row - 1, end_column=1)
        dc = ws.cell(day_start, 1, DAY_FULL[DAYS[d]])
        dc.fill = DAYFILL
        dc.font = Font(bold=True, color="FFFFFF", size=11)
        dc.alignment = CENTER_ROT
        for rr in range(day_start, row):
            ws.cell(rr, 1).fill = DAYFILL
            ws.cell(rr, 1).border = BORDER

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    for j in range(3, 3 + len(m.classes)):
        ws.column_dimensions[get_column_letter(j)].width = 16
    for r in range(2, row):
        ws.row_dimensions[r].height = 34
    ws.freeze_panes = "C2"
    return ws


def write_teacher_sheet(wb, m: Model, solution):
    cfg = m.cfg
    ws = wb.create_sheet(cfg.sheet_teacher_name)
    _cell(ws, 1, 1, cfg.teacher_title, NAVY, Font(bold=True, color="FFFFFF", size=13))
    ws.merge_cells("A1:I1")

    tgrid = defaultdict(lambda: [["" for _ in range(9)] for _ in range(N_DAYS)])
    for (c, d, p), (s, t) in solution.items():
        tgrid[t][d][p] = f"{cfg.class_display.get(c, c)} ({_abbr(cfg, s)})"
    for c in m.study_hour_classes:
        t = m.study_supervisor.get(c)
        if t:
            for d in range(N_DAYS):
                if m.has_p8(DAYS[d]):
                    tgrid[t][d][STUDY_PERIOD] = cfg.class_display.get(c, c)

    generic = [g for g in cfg.generic_teacher.values()
               if any(tt == g for _, tt in solution.values())]
    order = list(m.teachers) + generic

    row = 3
    for t in order:
        _cell(ws, row, 1, f"TEACHER: {t}", TEACHHDR, Font(bold=True, color="FFFFFF", size=11),
              Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        for j, h in enumerate(["DAY", "PERIOD 1", "PERIOD 2", "PERIOD 3", "PERIOD 4",
                               "PERIOD 5", "PERIOD 6", "PERIOD 7", "STUDY HOUR"], start=1):
            _cell(ws, row, j, h, DAYFILL, WHITE_BOLD)
        row += 1
        for d in range(N_DAYS):
            _cell(ws, row, 1, DAYS[d], PERFILL, Font(bold=True, size=10))
            for p in range(1, 9):
                txt = tgrid[t][d][p]
                col = 9 if p == STUDY_PERIOD else p + 1
                fill = STUDYFILL if (p == STUDY_PERIOD and txt) else None
                _cell(ws, row, col, txt, fill)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 10
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["I"].width = 15
    return ws


def write_workbook(path, m: Model, solution):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_teacher_sheet(wb, m, solution)
    write_class_sheet(wb, m, solution)
    wb.save(path)
    return path
