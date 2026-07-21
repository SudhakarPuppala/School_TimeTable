#!/usr/bin/env python3
"""Independent audit: re-parse a generated timetable .xlsx (both sheets) and
verify every class/period/teacher against the source workbook. School-aware.

Usage:  python audit.py --school NRCS \
                        --src NRCS/Requirements/NRCS_information.xlsx \
                        --out NRCS/Output/NRCS_Timetable_Final.xlsx
"""
import argparse
from collections import defaultdict

import openpyxl
from timetable.model import load_model, DAYS, STUDY_PERIOD, GENERIC_TEACHERS, INPUTS


def parse_class_sheet(path, m):
    abbr2subj = {m.abbr(s): s for s in m.subjects}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[m.cfg.sheet_class_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_class = {c: str(headers[c - 1]).strip()
                 for c in range(3, ws.max_column + 1) if headers[c - 1]}
    out, supervisors, day = {}, {}, -1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            day += 1
        plabel = ws.cell(r, 2).value
        is_study = str(plabel).strip().lower().startswith("study")
        period = STUDY_PERIOD if is_study else int(plabel)
        for c, cls in col_class.items():
            raw = ws.cell(r, c).value
            if not raw or str(raw).strip().lower() in ("recreation", "free period", "—", "-"):
                continue
            parts = str(raw).split("\n")
            teacher = parts[0].strip()
            abbr = parts[1].strip("() ").strip() if len(parts) > 1 else ""
            if abbr.upper() == "CLASS":
                supervisors[cls] = teacher
                continue
            out[(cls, day, period)] = (teacher, abbr2subj.get(abbr, abbr))
    return out, supervisors


def parse_teacher_sheet(path, m):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[m.cfg.sheet_teacher_name]
    grid, teacher = defaultdict(set), None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and str(a).startswith("TEACHER:"):
            teacher = str(a).split(":", 1)[1].strip()
            continue
        if a in DAYS and teacher:
            day = DAYS.index(a)
            for c in range(2, 10):
                raw = ws.cell(r, c).value
                if raw:
                    for line in str(raw).split("\n"):     # combined sessions: 1 class/line
                        if line.strip():
                            grid[(teacher, day, 8 if c == 9 else c - 1)].add(line.strip())
    return grid


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--school", default="NRHS", choices=["NRHS", "NRCS"])
    ap.add_argument("--src", help="source information workbook (default: standard path)")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    m = load_model(args.src or INPUTS[args.school], args.school)
    sol, supers = parse_class_sheet(args.out, m)
    tgrid = parse_teacher_sheet(args.out, m)
    fail = []

    # 1. weekly counts match the plan
    got = defaultdict(int)
    for (c, d, p), (t, s) in sol.items():
        got[(c, s)] += 1
    for c in m.classes:
        for s in m.subjects:
            want = m.plan.get((c, s), 0)
            if got[(c, s)] != want:
                fail.append(f"[COUNT] {c} {s}: plan={want} timetable={got[(c, s)]}")

    # 2. teacher matches the allotment
    for (c, d, p), (t, s) in sol.items():
        exp = m.teacher_of.get((c, s))
        if exp is None:
            fail.append(f"[NO-ALLOT] {c} {s} taught by {t}")
        elif exp not in GENERIC_TEACHERS and t != exp:
            fail.append(f"[WRONG-TEACHER] {c} {s} @ {DAYS[d]}P{p}: got {t}, allotment {exp}")

    # 3. no double-booking
    occ = defaultdict(list)
    for (c, d, p), (t, s) in sol.items():
        if t in GENERIC_TEACHERS or m.teacher_of.get((c, s)) in GENERIC_TEACHERS:
            continue
        occ[(t, d, p)].append(c)
    for (t, d, p), cs in occ.items():
        if len(cs) > 1:
            fail.append(f"[DOUBLE-BOOK] {t} @ {DAYS[d]}P{p}: {cs}")

    # 4. packing: taught slots == planned content, inside real slots
    for c in m.classes:
        content = sum(m.plan.get((c, s), 0) for s in m.subjects)
        if c in m.study_hour_classes:
            filled = sum(1 for d in range(6) for p in range(1, 8) if (c, d, p) in sol)
            if filled != content:
                fail.append(f"[PACK] {c}: {filled} taught in P1-P7 but content is {content}")
            extra = [p for d in range(6) for p in (8,) if (c, d, p) in sol]
            if extra:
                fail.append(f"[P8] {c} has taught periods at P8 but a Study Hour exists")
        else:
            filled = sum(1 for d in range(6) for p in range(1, 9) if (c, d, p) in sol)
            if filled != content:
                fail.append(f"[PACK] {c}: {filled} taught but content is {content}")
    for (c, d, p), (t, s) in sol.items():
        if p == 8 and DAYS[d] in m.cfg.no_p8_days:
            fail.append(f"[NO-P8-DAY] {c} scheduled at {DAYS[d]} P8 (no P8 that day)")

    # 5. study-hour supervisors match
    for c in m.study_hour_classes:
        if supers.get(c) != m.study_supervisor.get(c):
            fail.append(f"[SUPERVISOR] {c}: sheet={supers.get(c)} source={m.study_supervisor.get(c)}")

    # 6. Teacher Leisure Plan honoured (MUST = hard)
    for (c, d, p), (t, s) in sol.items():
        if t in GENERIC_TEACHERS:
            continue
        if p in m.blocked.get(t, set()):
            fail.append(f"[LEISURE] {t} @ {DAYS[d]}P{p} ({c}) is marked Leisure (MUST)")

    # 6b. Activity-Plan windows honoured (days × periods, union over rows)
    for (c, d, p), (t, s) in sol.items():
        if m.has_activity_window(s, c) and not m.activity_allows(s, c, d, p):
            fail.append(f"[ACTIVITY-WINDOW] {s} ({c}) @ {DAYS[d]}P{p} is outside its "
                        f"Activity-Plan window(s)")

    # 7. Class sheet and Teacher sheet tally BOTH WAYS, generic instructors
    #    (P.E / MARTIAL ARTS combined sessions) included: every entry on either
    #    sheet must appear on the other.
    expected = set()
    for (c, d, p), (t, s) in sol.items():
        expected.add((t, d, p, f"{c} ({m.abbr(s)})"))
    for c in m.study_hour_classes:
        t = m.study_supervisor.get(c)
        for d in range(6):
            if t and m.has_p8(DAYS[d]):
                expected.add((t, d, 8, c))
    parsed = {(t, d, p, line) for (t, d, p), lines in tgrid.items() for line in lines}
    for t, d, p, line in sorted(expected - parsed):
        fail.append(f"[SHEET-MISSING] {t} @ {DAYS[d]}P{p}: Class sheet has '{line}' "
                    f"but the Teacher sheet does not")
    for t, d, p, line in sorted(parsed - expected):
        fail.append(f"[SHEET-EXTRA] {t} @ {DAYS[d]}P{p}: Teacher sheet shows '{line}' "
                    f"but the Class sheet does not")

    print(f"INDEPENDENT AUDIT of {args.out}  (school={args.school})")
    print("=" * 60)
    for ck in ["weekly counts", "teacher==allotment", "no double-booking",
               "grid packed / free-slots", "study-hour supervisors",
               "Teacher Leisure Plan (MUST)", "Activity-Plan windows",
               "Class<->Teacher sheet agreement (both ways, incl. combined sessions)"]:
        print(f"  + checked: {ck}")
    print("=" * 60)
    if fail:
        print(f"X {len(fail)} problem(s):")
        for f in fail:
            print("   ", f)
    else:
        print(f"OK ALL CHECKS PASSED - {len(sol)} taught slots across "
              f"{len(m.classes)} classes and {len(m.teachers)} teachers.")
    return 1 if fail else 0


if __name__ == "__main__":
    raise SystemExit(main())
