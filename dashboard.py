"""Streamlit dashboard for the school timetable generator.

Run:  streamlit run dashboard.py

Edit the four source sheets (including the Teacher Leisure Plan) in the
browser, check conflicts (offending cells highlighted in RED), re-solve with
OR-Tools, review the colour-coded Class & Teacher timetables, and download
the styled workbook / PDF.
"""
from __future__ import annotations
import os
import re
import subprocess
import tempfile
from collections import defaultdict
from datetime import datetime

import numpy as np
import pandas as pd
import openpyxl
import streamlit as st

from timetable.model import (load_model, DAYS, GENERIC_TEACHERS, _is_tick,
                             INPUTS, LEISURE_COLS, SHEET_PLAN, SHEET_ALLOT,
                             SHEET_P1, SHEET_LEISURE, SHEET_ACTIVITY,
                             SCHOOLS as SCHOOLS_CFG)
from timetable.conflicts import (check_conflicts, has_errors, error_cells,
                                 class_capacity, class_content,
                                 teacher_capacity, teacher_demand)
from timetable.solver import solve
from timetable.verify import verify
from timetable.writer import write_workbook
from timetable.pdf import write_pdf

PLABEL = ["P1", "P2", "P3", "P4", "P5", "P6", "P7", "Study Hour"]
SUBJ_COLOR = {
    "TEL": ("#E6F1FB", "#0C447C"), "HIN": ("#EEEDFE", "#3C3489"),
    "ENG": ("#E1F5EE", "#085041"), "GRAM": ("#D6EFE6", "#0F6E56"), "MATH": ("#FAEEDA", "#633806"),
    "EVS": ("#EAF3DE", "#27500A"), "PHY": ("#FAECE7", "#712B13"),
    "CHEM": ("#FBEAF0", "#72243E"), "BIO": ("#C0DD97", "#173404"),
    "SOC": ("#F1EFE8", "#2C2C2A"), "COMP": ("#B5D4F4", "#042C53"),
    "P.E.T": ("#FCEBEB", "#791F1F"), "G.K": ("#FAC775", "#412402"),
    "ORAL": ("#F4C0D1", "#4B1528"), "KAR": ("#F5C4B3", "#4A1B0C"),
    "STUDY": ("#D3D1C7", "#2C2C2A"),
}
CLASS_PALETTE = [
    ("#E6F1FB", "#0C447C"), ("#E1F5EE", "#085041"), ("#FAEEDA", "#633806"),
    ("#EEEDFE", "#3C3489"), ("#FBEAF0", "#72243E"), ("#EAF3DE", "#27500A"),
    ("#FAECE7", "#712B13"), ("#E0F2F1", "#00473E"), ("#FFF3E0", "#7A4A00"),
    ("#EDE7F6", "#3A2A6A"), ("#FCE4EC", "#761B39"), ("#E8F5E9", "#1B4D22"),
    ("#E3F2FD", "#0B3D66"), ("#F9FBE7", "#4A5210"), ("#EFEBE9", "#4A342E"),
]
RED_CELL = "background-color:#B3261E;color:#FFFFFF;font-weight:600"

st.set_page_config(page_title="School Timetable Generator", layout="wide")


# ------------------------------------------------------------------ data I/O
def _fill_count(row):
    return sum(1 for v in row[1:] if v not in (None, ""))


def read_frames(path):
    """Read the 4 source sheets into editable DataFrames (layout-sniffing)."""
    wb = openpyxl.load_workbook(path, data_only=True)

    def rows_of(name):
        for ws in wb.worksheets:
            if ws.title.strip().lower() == name.lower():
                return [list(r) for r in ws.iter_rows(values_only=True)]
        return []

    # --- Weekly Period Plan ---
    rows = rows_of(SHEET_PLAN)
    hdr = next((i for i, r in enumerate(rows) if _fill_count(r) >= 3), 0)
    cols = [j for j, v in enumerate(rows[hdr]) if j >= 1 and v not in (None, "")]
    classes = [str(rows[hdr][j]).strip() for j in cols]
    body = []
    for r in rows[hdr + 1:]:
        if r[0] in (None, "") or str(r[0]).strip().lower().startswith("total"):
            continue
        body.append([str(r[0]).strip()] +
                    [int(r[j]) if j < len(r) and r[j] not in (None, "") else 0 for j in cols])
    plan = pd.DataFrame(body, columns=["Subject"] + classes)

    # --- Teacher Allotment ---
    rows = rows_of(SHEET_ALLOT)
    hdr = next((i for i, r in enumerate(rows) if _fill_count(r) >= 3), 0)
    acols = [j for j, v in enumerate(rows[hdr]) if j >= 1 and v not in (None, "")]
    aclasses = [str(rows[hdr][j]).strip() for j in acols]
    body = []
    for r in rows[hdr + 1:]:
        if r[0] in (None, ""):
            continue
        body.append([str(r[0]).strip()] +
                    [("" if j >= len(r) or r[j] in (None, "") else str(r[j]).strip())
                     for j in acols])
    allot = pd.DataFrame(body, columns=["Subject"] + aclasses)

    # --- Period 1 teacher allotment (both layouts) ---
    rows = rows_of(SHEET_P1)
    p1_map, sh_map, pclasses = {}, {}, []
    day_i = next((i for i, r in enumerate(rows)
                  if r and str(r[0]).strip().lower() == "day"), None)
    if day_i is not None:
        pclasses = [str(v).strip() for v in rows[day_i][2:] if v not in (None, "")]
        idx = [j for j, v in enumerate(rows[day_i]) if j >= 2 and v not in (None, "")]
        for r in rows[day_i + 1:]:
            lbl = str(r[1]).strip().lower() if len(r) > 1 and r[1] is not None else ""
            tgt = p1_map if lbl in ("1", "1.0") else (sh_map if lbl.startswith("study") else None)
            if tgt is None:
                continue
            for cl, j in zip(pclasses, idx):
                if j < len(r) and r[j] not in (None, ""):
                    tgt[cl] = str(r[j]).strip()
    else:
        hdr = next((i for i, r in enumerate(rows) if _fill_count(r) >= 3), 0)
        idx = [j for j, v in enumerate(rows[hdr]) if j >= 1 and v not in (None, "")]
        pclasses = [str(rows[hdr][j]).strip() for j in idx]
        for r in rows[hdr + 1:]:
            lbl = str(r[0]).strip().lower() if r and r[0] is not None else ""
            tgt = (p1_map if lbl.startswith("period 1") else
                   (sh_map if lbl.startswith("study") else None))
            if tgt is None:
                continue
            for cl, j in zip(pclasses, idx):
                if j < len(r) and r[j] not in (None, ""):
                    tgt[cl] = str(r[j]).strip()
    if not pclasses:
        pclasses = classes
    p1 = pd.DataFrame({"Class": pclasses,
                       "Period 1 Teacher": [p1_map.get(c, "") for c in pclasses],
                       "Study Hour": [sh_map.get(c, "") for c in pclasses]})

    # --- Teacher Leisure Plan ---
    rows = rows_of(SHEET_LEISURE)
    hdr = next((i for i, r in enumerate(rows)
                if r and r[0] and str(r[0]).strip().lower().startswith("teacher name")), None)
    body = []
    if hdr is not None:
        # map columns by header text
        col_of = {}
        for j, v in enumerate(rows[hdr]):
            if v in (None, ""):
                continue
            k = str(v).strip().lower()
            if "fitm" in k:
                col_of["Leisure Fitment"] = j
            elif k.startswith("period"):
                col_of[f"Period {''.join(ch for ch in k if ch.isdigit())}"] = j
            elif k.startswith("study"):
                col_of["Study Hour"] = j
        for r in rows[hdr + 1:]:
            if r[0] in (None, ""):
                continue
            rec = {"Teacher Name": str(r[0]).strip().upper().replace(",", "."),
                   "Lunch Break": "Lunch Break"}
            for name, j in col_of.items():
                v = r[j] if j < len(r) else None
                rec[name] = "" if v in (None, "") else str(v).strip()
            body.append(rec)
    leisure = pd.DataFrame(body, columns=LEISURE_COLS)
    if not body:
        leisure = pd.DataFrame(columns=LEISURE_COLS)

    # --- Activity Plan (optional sheet; each row = one combined-session group) ---
    rows = rows_of(SHEET_ACTIVITY)
    act_cols = ["Activity", "Allowed Days", "Allowed Periods"] + classes
    body = []
    hdr = next((i for i, r in enumerate(rows)
                if r and r[0] and str(r[0]).strip().lower().startswith("activity")
                and _fill_count(r) >= 1), None)
    if hdr is not None:
        head = [str(v).strip() if v not in (None, "") else "" for v in rows[hdr]]
        for r in rows[hdr + 1:]:
            if not r or r[0] in (None, ""):
                continue
            rec = {c: (False if c in classes else "") for c in act_cols}
            rec["Activity"] = str(r[0]).strip()
            for j, h in enumerate(head):
                if j == 0 or j >= len(r) or r[j] in (None, ""):
                    continue
                hl = h.lower()
                if "day" in hl and "period" not in hl:
                    rec["Allowed Days"] = str(r[j]).strip()
                elif "period" in hl or hl == "allowed":
                    rec["Allowed Periods"] = str(r[j]).strip()
                elif h in classes:
                    rec[h] = _is_tick(r[j])
            body.append(rec)
    activity = pd.DataFrame(body, columns=act_cols)
    for c in classes:
        if c in activity.columns:
            activity[c] = activity[c].fillna(False).astype(bool)
    return {"plan": plan, "allot": allot, "p1": p1, "leisure": leisure,
            "activity": activity}


def _cell(v):
    return "" if v is None or (isinstance(v, float) and np.isnan(v)) or pd.isna(v) else v


def write_frames(frames, path):
    """Write edited DataFrames back in the canonical 4-sheet layout."""
    plan, allot, p1, leisure = (frames["plan"], frames["allot"],
                                frames["p1"], frames["leisure"])
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(SHEET_PLAN)
    ws.append([SHEET_PLAN])
    ws.append(list(plan.columns))
    for _, r in plan.iterrows():
        ws.append([r.iloc[0]] + [int(v) if _cell(v) != "" else 0 for v in r.iloc[1:]])
    ws.append(["Total"] + [int(plan[c].fillna(0).sum()) for c in plan.columns[1:]])

    ws = wb.create_sheet(SHEET_ALLOT)
    ws.append([SHEET_ALLOT])
    ws.append(list(allot.columns))
    for _, r in allot.iterrows():
        ws.append([_cell(v) for v in r.tolist()])

    ws = wb.create_sheet(SHEET_P1)
    ws.append(["Period 1 Teacher Allotment"])
    ws.append(["Class"] + p1["Class"].tolist())
    ws.append(["Period 1 Teacher"] + [_cell(v) for v in p1["Period 1 Teacher"].tolist()])
    ws.append(["Study Hour"] + [_cell(v) for v in p1["Study Hour"].tolist()])

    ws = wb.create_sheet(SHEET_LEISURE)
    ws.append([SHEET_LEISURE])
    ws.append(LEISURE_COLS)
    for _, r in leisure.iterrows():
        if _cell(r.get("Teacher Name", "")) == "":
            continue
        ws.append([_cell(r.get(c, "")) if c != "Lunch Break" else "Lunch Break"
                   for c in LEISURE_COLS])

    activity = frames.get("activity")
    if activity is not None:
        ws = wb.create_sheet(SHEET_ACTIVITY)
        ws.append([SHEET_ACTIVITY + "  (each row = one combined session group)"])
        ws.append(list(activity.columns))
        for _, r in activity.iterrows():
            if _cell(r.get("Activity", "")) == "":
                continue
            row = []
            for col in activity.columns:
                v = r[col]
                if col in ("Activity", "Allowed Days", "Allowed Periods"):
                    row.append(_cell(v))
                else:
                    row.append("Yes" if v is True or _is_tick(v) else "")
            ws.append(row)
    wb.save(path)
    return path


def frames_to_model(frames, school):
    tmp = os.path.join(tempfile.gettempdir(), f"tt_edit_{school}.xlsx")
    write_frames(frames, tmp)
    return load_model(tmp, school), tmp


def _github_token():
    """GitHub token from Streamlit secrets (for cloud deployments)."""
    try:
        if "GITHUB_TOKEN" in st.secrets:
            return str(st.secrets["GITHUB_TOKEN"]).strip() or None
        if "github" in st.secrets and "token" in st.secrets["github"]:
            return str(st.secrets["github"]["token"]).strip() or None
    except Exception:
        pass
    return None


def git_commit_push(path, school):
    """Commit the saved workbook and push to GitHub. -> (ok, message).

    Works both locally (uses your normal git credentials) and on Streamlit
    Cloud: there the container has no git identity or credentials, so the
    commit uses a built-in identity and the push authenticates with a GitHub
    token read from the app's Secrets (GITHUB_TOKEN = "...").
    """
    repo = os.path.dirname(os.path.abspath(__file__))
    token = _github_token()

    def run(*args):
        return subprocess.run(["git", *args], cwd=repo, capture_output=True,
                              text=True, timeout=120)

    def clean(*texts):
        out = " ".join(t.strip() for t in texts if t and t.strip())
        return out.replace(token, "***") if token else out

    r = run("add", "--", os.path.abspath(path))
    if r.returncode != 0:
        return False, f"git add failed: {clean(r.stderr, r.stdout)}"
    r = run("diff", "--cached", "--quiet", "--", os.path.abspath(path))
    if r.returncode == 0:
        return True, "No changes to commit (workbook identical to last commit)."

    ident = []
    if not (run("config", "user.email").stdout or "").strip():
        ident = ["-c", "user.name=Timetable Dashboard",
                 "-c", "user.email=timetable-dashboard@users.noreply.github.com"]
    stamp = datetime.now().strftime("%d %b %Y %H:%M")
    r = run(*ident, "commit", "-m",
            f"{school}: dashboard edit of information workbook ({stamp})")
    if r.returncode != 0:
        return False, f"git commit failed: {clean(r.stderr, r.stdout)}"

    branch = (run("rev-parse", "--abbrev-ref", "HEAD").stdout or "").strip()
    if branch in ("", "HEAD"):
        branch = "main"
    push_args, auth_url = ["push"], None
    if token:
        origin = (run("remote", "get-url", "origin").stdout or "").strip()
        mrepo = re.search(r"github\.com[:/](.+?)(?:\.git)?$", origin)
        if mrepo:
            auth_url = f"https://x-access-token:{token}@github.com/{mrepo.group(1)}.git"
            push_args = ["push", auth_url, f"HEAD:{branch}"]
    r = run(*push_args)
    if r.returncode != 0 and auth_url and re.search(
            r"fetch first|rejected|non-fast-forward", r.stderr or ""):
        pr = run(*ident, "pull", "--rebase", auth_url, branch)
        if pr.returncode == 0:
            r = run(*push_args)
        else:
            run("rebase", "--abort")
    if r.returncode != 0:
        hint = ("" if token else
                " — on Streamlit Cloud, add a GitHub token to the app's Secrets as "
                'GITHUB_TOKEN = "..." so the dashboard can push (see README)')
        return False, (f"Committed in the app, but git push failed{hint}: "
                       f"{clean(r.stderr)[:300]}")
    extra = (" The cloud app will restart with the new data in a minute."
             if token else "")
    return True, "Committed and pushed to GitHub." + extra


# ------------------------------------------------------------------ visual grids
def class_grid(m, solution, cls):
    g = [["—"] * 6 for _ in range(8)]
    for d in range(6):
        for p in range(1, 9):
            if (cls, d, p) in solution:
                s, t = solution[(cls, d, p)]
                g[p - 1][d] = (m.abbr(s), t, "subj")
            elif p == 8 and cls in m.study_hour_classes and m.has_p8(DAYS[d]):
                g[p - 1][d] = ("STUDY", m.study_supervisor.get(cls, ""), "subj")
    return g


def teacher_grid(m, solution, teacher, class_color):
    """Cells hold every class in the slot — a combined P.E/Karate session has
    several classes at once."""
    cell = defaultdict(list)
    pos = {c: i for i, c in enumerate(m.classes)}
    for (c, d, p), (s, t) in solution.items():
        if t == teacher:
            cell[(d, p)].append((pos.get(c, 99), c, m.abbr(s)))
    for c in m.study_hour_classes:
        if m.study_supervisor.get(c) == teacher:
            for d in range(6):
                if m.has_p8(DAYS[d]):
                    cell[(d, 8)].append((pos.get(c, 99), c, "STUDY"))
    g = [["—"] * 6 for _ in range(8)]
    for (d, p), entries in cell.items():
        entries.sort()
        if len(entries) == 1:
            _, c, ab = entries[0]
            g[p - 1][d] = (c, ab, "class")
        else:                                   # combined session
            ab = entries[0][2]
            g[p - 1][d] = (ab, ", ".join(c for _, c, _ in entries), "subj")
    return g


def grid_html(g, class_color=None):
    rows = ["<table style='width:100%;border-collapse:separate;border-spacing:3px;table-layout:fixed'>"]
    rows.append("<tr><th style='width:60px'></th>" +
                "".join(f"<th style='font-size:12px;color:#666;padding:4px'>{d}</th>" for d in DAYS) + "</tr>")
    for p in range(8):
        cells = [f"<td style='font-size:12px;color:#666;text-align:right;padding-right:6px'>{PLABEL[p]}</td>"]
        for d in range(6):
            raw = g[p][d]
            if raw == "—":
                cells.append("<td style='background:#f4f4f2;border-radius:6px;height:44px'></td>")
                continue
            top, bot, mode = raw
            if mode == "class" and class_color:
                bg, fg = class_color.get(top, ("#F1EFE8", "#2C2C2A"))
            else:
                bg, fg = SUBJ_COLOR.get(top, ("#F1EFE8", "#2C2C2A"))
            cells.append(
                f"<td style='background:{bg};color:{fg};border-radius:6px;height:44px;"
                f"text-align:center;vertical-align:middle;padding:3px'>"
                f"<div style='font-size:12px;font-weight:600;line-height:1.1'>{top}</div>"
                f"<div style='font-size:10px;opacity:.85;line-height:1.15'>{bot}</div></td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    rows.append("</table>")
    return "".join(rows)


def legend_html(pairs):
    chips = "".join(
        f"<span style='background:{bg};color:{fg};border-radius:6px;padding:2px 10px;"
        f"margin:2px;font-size:11px;font-weight:600;display:inline-block'>{label}</span>"
        for label, (bg, fg) in pairs)
    return f"<div style='margin:4px 0 10px'>{chips}</div>"


def _section(title, body_html):
    st.markdown(
        f"<div style='background:#2A4D69;color:#fff;font-weight:600;font-size:16px;"
        f"padding:8px 14px;border-radius:8px;margin:18px 0 8px'>{title}</div>",
        unsafe_allow_html=True)
    st.markdown(body_html, unsafe_allow_html=True)


# ------------------------------------------------------------------ conflict views
def style_red(df, key_col, bad):
    """Return a Styler with RED on cells listed in bad = {(row_key, col)}."""
    def apply(row):
        rk = str(row[key_col]).strip()
        return [RED_CELL if (rk, str(col)) in bad or (rk, "") in bad else ""
                for col in row.index]
    return df.style.apply(apply, axis=1)


def show_conflicts(conflicts):
    errs = [c for c in conflicts if c.severity == "error"]
    warns = [c for c in conflicts if c.severity == "warning"]
    infos = [c for c in conflicts if c.severity == "info"]
    if errs:
        st.error("**" + f"{len(errs)} conflict(s) must be fixed before generating:**\n\n" +
                 "\n".join(f"- **[{c.sheet}]** {c.message}" for c in errs))
    else:
        st.success("No blocking conflicts — you can generate the timetable.")
    if warns:
        with st.expander(f"⚠️ Warnings ({len(warns)})", expanded=bool(errs)):
            for c in warns:
                st.markdown(f"- **[{c.sheet}]** {c.message}")
    if infos:
        with st.expander(f"ℹ️ Notes — auto-fills & name matching ({len(infos)})"):
            for c in infos:
                st.markdown(f"- {c.message}")


# ------------------------------------------------------------------ UI
st.title("🗓️  School Timetable Generator")
st.caption("Rules live in the **Teacher Leisure Plan** sheet — edit data, check "
           "conflicts, generate, download. Powered by OR-Tools CP-SAT.")

_qp_school = st.query_params.get("school", "").upper()
_default_school = _qp_school if _qp_school in INPUTS else "NRHS"

with st.sidebar:
    st.header("Setup")
    school = st.selectbox("School", list(INPUTS.keys()),
                          index=list(INPUTS.keys()).index(_default_school))
    _default_path = st.query_params.get("path", "") if _qp_school == school else ""
    input_path = st.text_input("Information workbook", _default_path or INPUTS[school])
    seconds = st.slider("Solver time budget (s)", 15, 240, 90, 15)
    st.divider()
    check_btn = st.button("🔍  Check conflicts", width="stretch")
    generate = st.button("⚙️  Generate timetable", type="primary", width="stretch")
    save_src = st.button("💾  Save edits to source workbook", width="stretch")
    auto_git = st.checkbox("⬆️ Auto commit & push to GitHub on save", value=True)
    if auto_git:
        if _github_token():
            st.caption("🔑 GitHub token detected in Secrets — pushes will use it.")
        else:
            st.caption("🔓 No GitHub token in Secrets — pushes use local git "
                       "credentials (fine on your PC; on Streamlit Cloud add "
                       '`GITHUB_TOKEN = "..."` under Manage app → Settings → Secrets).')
    st.caption("Check = validate the current edits and highlight conflicts in red. "
               "Generate = conflict-check + solve. Save = write edits back to the workbook"
               " (and check them into GitHub when the box is ticked).")

if (st.session_state.get("school"), st.session_state.get("path")) != (school, input_path):
    st.session_state.school = school
    st.session_state.path = input_path
    for k in ("frames", "result", "conflicts", "bad_cells", "gen_feedback"):
        st.session_state.pop(k, None)
if "frames" not in st.session_state and os.path.exists(input_path):
    st.session_state.frames = read_frames(input_path)

# prominent main-area banner for the last Generate attempt (filled by the
# Generate handler further down; shown here at the top so it can't be missed)
status_area = st.container()

tab_data, tab_class, tab_teacher, tab_report = st.tabs(
    ["✏️ Edit data & conflicts", "📚 Class timetable", "👩‍🏫 Teacher timetable",
     "✅ Validation & report"])

bad = st.session_state.get("bad_cells", {})

with tab_data:
    if "frames" not in st.session_state:
        st.warning(f"Workbook not found: {input_path}")
    else:
        f = st.session_state.frames
        edited = {}

        # ---------- Weekly Period Plan + totals ----------
        st.subheader("Weekly period plan")
        num_cols = {c: st.column_config.NumberColumn(c, min_value=0, max_value=48, step=1)
                    for c in f["plan"].columns[1:]}
        edited["plan"] = st.data_editor(f["plan"], width="stretch", num_rows="dynamic",
                                        key=f"plan_{school}", column_config=num_cols)
        plan_e = edited["plan"]
        # totals row, like Excel — live, with capacity per class
        totals = {c: int(pd.to_numeric(plan_e[c], errors="coerce").fillna(0).sum())
                  for c in plan_e.columns[1:]}
        sh_col = {str(r["Class"]).strip(): str(_cell(r["Study Hour"])).strip()
                  for _, r in f["p1"].iterrows()}
        p8_days = 6 - len(SCHOOLS_CFG[school].no_p8_days)
        caps = {c: 42 if sh_col.get(c, "") else 42 + p8_days for c in plan_e.columns[1:]}
        tot_df = pd.DataFrame(
            [["Total (planned)"] + [totals[c] for c in plan_e.columns[1:]],
             ["Capacity (slots)"] + [caps[c] for c in plan_e.columns[1:]]],
            columns=list(plan_e.columns))

        def tot_style(row):
            out = [""]
            for c in row.index[1:]:
                if row.iloc[0] == "Capacity (slots)":
                    out.append("color:#666")
                elif totals[c] > caps[c]:
                    out.append(RED_CELL)
                else:
                    out.append("background-color:#0F6E56;color:white;font-weight:600")
            return out
        st.dataframe(tot_df.style.apply(tot_style, axis=1), width="stretch", hide_index=True)
        st.caption("Green = within capacity · **Red = over-assigned**. A class **with** a "
                   "Study-Hour supervisor has 42 teachable slots (P1-P7 × 6 days); without "
                   f"one, Period 8 is teachable too (+{p8_days}).")
        if SHEET_PLAN in bad or SHEET_ALLOT in bad:
            pb = bad.get(SHEET_PLAN, set()) | bad.get(SHEET_ALLOT, set())
            with st.expander("🔴 Plan cells named in conflicts", expanded=True):
                st.dataframe(style_red(plan_e, "Subject", pb), width="stretch", hide_index=True)

        # ---------- Teacher Allotment ----------
        st.subheader("Teacher allotment")
        edited["allot"] = st.data_editor(f["allot"], width="stretch", num_rows="dynamic",
                                         key=f"allot_{school}")
        if SHEET_ALLOT in bad:
            with st.expander("🔴 Allotment cells named in conflicts", expanded=True):
                st.dataframe(style_red(edited["allot"], "Subject", bad[SHEET_ALLOT]),
                             width="stretch", hide_index=True)

        # ---------- Period 1 / Study hour ----------
        st.subheader("Period 1 teacher & Study-Hour supervisor")
        edited["p1"] = st.data_editor(f["p1"], width="stretch", key=f"p1_{school}")
        if SHEET_P1 in bad:
            with st.expander("🔴 Period-1 cells named in conflicts", expanded=True):
                st.dataframe(style_red(edited["p1"], "Class", bad[SHEET_P1]),
                             width="stretch", hide_index=True)

        # ---------- Teacher Leisure Plan ----------
        st.subheader("Teacher Leisure Plan")
        st.caption("**MUST** = leisure is strict (never scheduled there). **BEST** = the "
                   "solver keeps gaps between continuous periods where possible. "
                   "Mark a period with **Leisure** to reserve it; Lunch Break is free for everyone.")
        leisure_cfg = {"Leisure Fitment": st.column_config.SelectboxColumn(
                           "Leisure Fitment", options=["MUST", "BEST"], required=True),
                       "Lunch Break": st.column_config.TextColumn("Lunch Break", disabled=True)}
        for c in LEISURE_COLS:
            if c.startswith("Period") or c == "Study Hour":
                leisure_cfg[c] = st.column_config.SelectboxColumn(c, options=["", "Leisure"])
        edited["leisure"] = st.data_editor(f["leisure"], width="stretch", num_rows="dynamic",
                                           key=f"leisure_{school}", column_config=leisure_cfg)
        if SHEET_LEISURE in bad:
            with st.expander("🔴 Leisure-Plan cells named in conflicts", expanded=True):
                st.dataframe(style_red(edited["leisure"], "Teacher Name", bad[SHEET_LEISURE]),
                             width="stretch", hide_index=True)

        # ---------- Activity Plan (P.E.T / Karate) ----------
        st.subheader("Activity Plan — parallel activities (P.E.T, Karate)")
        st.caption("**Each row is one combined session group**: tick the classes that do "
                   "the activity **together**, and set that row's **Allowed Days** "
                   "(e.g. `MON,TUE` or `MON-THU` — blank = any day) and **Allowed "
                   "Periods** (e.g. `6,7` — blank = any period). Add rows to make more "
                   "combinations. A class ticked in no row is scheduled independently "
                   "with no restriction.")
        act_cfg = {"Activity": st.column_config.TextColumn("Activity"),
                   "Allowed Days": st.column_config.TextColumn(
                       "Allowed Days", help="e.g. MON,TUE or MON-THU or blank for any day"),
                   "Allowed Periods": st.column_config.TextColumn(
                       "Allowed Periods", help="e.g. 6,7 or 5-7 or blank for any period")}
        for c in f["activity"].columns[3:]:
            act_cfg[c] = st.column_config.CheckboxColumn(c, default=False)
        edited["activity"] = st.data_editor(f["activity"], width="stretch",
                                            num_rows="dynamic", key=f"activity_{school}",
                                            column_config=act_cfg)
        if SHEET_ACTIVITY in bad:
            with st.expander("🔴 Activity-Plan cells named in conflicts", expanded=True):
                st.dataframe(style_red(edited["activity"], "Activity", bad[SHEET_ACTIVITY]),
                             width="stretch", hide_index=True)

        st.session_state.edited = edited

        # ---------- teacher load (over-assignment) ----------
        st.subheader("Teacher weekly load vs availability")
        try:
            m_live, _ = frames_to_model(edited, school)
            sup_days = m_live.study_days()
            recs = []
            for t in m_live.teachers:
                demand = teacher_demand(m_live, t)
                cap = teacher_capacity(m_live, t)
                sup = sum(sup_days for c in m_live.study_hour_classes
                          if m_live.study_supervisor.get(c) == t)
                recs.append({"Teacher": t, "Fitment": m_live.fitment.get(t, "BEST"),
                             "Teaching": demand, "Study-Hour supervision": sup,
                             "Total": demand + sup, "Teachable slots": cap,
                             "Spare": cap - demand})
            ldf = pd.DataFrame(recs).sort_values("Spare").reset_index(drop=True)

            def load_style(row):
                s = RED_CELL if row["Spare"] < 0 else ""
                return [s] * len(row.index)
            st.dataframe(ldf.style.apply(load_style, axis=1), width="stretch",
                         hide_index=True, height=min(38 * len(ldf) + 40, 600))
            st.caption("Red row = the teacher is allotted more periods than the Leisure "
                       "Plan allows. Spare = teachable slots − teaching load.")
        except Exception as e:
            st.warning(f"Load table unavailable: {e}")

        # ---------- conflict summary ----------
        if "conflicts" in st.session_state:
            st.divider()
            st.subheader("Conflict check result")
            show_conflicts(st.session_state.conflicts)


def run_conflict_check():
    m, _ = frames_to_model(st.session_state.edited, school)
    conflicts = check_conflicts(m)
    st.session_state.conflicts = conflicts
    st.session_state.bad_cells = error_cells(conflicts)
    return m, conflicts


if check_btn and "edited" in st.session_state:
    run_conflict_check()
    st.rerun()

if save_src and "edited" in st.session_state:
    try:
        write_frames(st.session_state.edited, input_path)
    except PermissionError:
        st.sidebar.error(f"Cannot save: {input_path} is open in Excel — close it and retry.")
        st.stop()
    msg = f"Saved edits to {input_path}"
    if auto_git:
        ok, gitmsg = git_commit_push(input_path, school)
        (st.sidebar.success if ok else st.sidebar.warning)(f"{msg}\n\n{gitmsg}")
    else:
        st.sidebar.success(msg)
    st.session_state.pop("frames", None)

if generate and "edited" in st.session_state:
    m, conflicts = run_conflict_check()
    errs = [c for c in conflicts if c.severity == "error"]
    if errs:
        st.session_state.pop("result", None)          # clear any stale timetable
        st.session_state.gen_feedback = ("blocked",
            f"{len(errs)} conflict(s) block generation — fix these first "
            f"(highlighted red in **✏️ Edit data & conflicts**):",
            [c.message for c in errs])
        st.sidebar.error(f"{len(errs)} conflict(s) — cannot generate. See the "
                         f"banner and the Edit tab.")
    else:
        try:
            with st.spinner("Solving… (this can take up to the time budget)"):
                solution, status, obj, notes = solve(m, max_seconds=seconds, precheck=False)
                errors, warnings = verify(m, solution)
                out = os.path.join(tempfile.gettempdir(), f"{school}_Timetable.xlsx")
                write_workbook(out, m, solution)
                pdf_out = os.path.join(tempfile.gettempdir(), f"{school}_Timetable.pdf")
                write_pdf(pdf_out, m, solution)
            st.session_state.result = (m, solution, status, obj, errors, warnings + notes)
            st.session_state.out = out
            st.session_state.pdf = pdf_out
            st.session_state.gen_feedback = ("success",
                f"Solved: {status} — {len(errors)} hard error(s), {len(warnings)} "
                f"warning(s). See the 📚 Class / 👩‍🏫 Teacher / ✅ Report tabs.", None)
            st.sidebar.success(f"Solved: {status}")
        except Exception as e:
            st.session_state.pop("result", None)
            st.session_state.gen_feedback = ("failed",
                f"Generation failed: {e}", None)
            st.sidebar.error("Generation failed — see the banner.")
    st.rerun()          # repaint: banner, red cells (Edit tab), result tabs

# render the Generate banner in the reserved top-of-page area
_fb = st.session_state.get("gen_feedback")
if _fb:
    kind, text, items = _fb
    with status_area:
        if kind == "success":
            st.success("✅ " + text)
        elif kind == "blocked":
            st.error("🚫 " + text + "\n\n" +
                     "\n".join(f"- {i}" for i in items))
        else:
            st.error("❌ " + text)

res = st.session_state.get("result")
if res:
    class_color = {c: CLASS_PALETTE[i % len(CLASS_PALETTE)]
                   for i, c in enumerate(res[0].classes)}

with tab_class:
    if res:
        m, solution = res[0], res[1]
        st.caption(f"{m.cfg.name} — all {len(m.classes)} class timetables (colour = subject)")
        used = sorted({m.abbr(s) for (c, d, p), (s, t) in solution.items()} | {"STUDY"})
        st.markdown(legend_html([(a, SUBJ_COLOR.get(a, ("#F1EFE8", "#2C2C2A")))
                                 for a in used]), unsafe_allow_html=True)
        for c in m.classes:
            _section(f"📚  {c}", grid_html(class_grid(m, solution, c)))
    else:
        st.info("Click **Generate timetable** to view results.")

with tab_teacher:
    if res:
        m, solution = res[0], res[1]
        teachers = sorted(m.teachers) + sorted(
            g for g in GENERIC_TEACHERS if any(t == g for _, t in solution.values()))
        st.caption(f"{m.cfg.name} — all {len(teachers)} teacher timetables (colour = class)")
        st.markdown(legend_html([(c, class_color[c]) for c in m.classes]),
                    unsafe_allow_html=True)
        for who in teachers:
            fit = m.fitment.get(who)
            tag = f" · {fit} leisure" if fit else ""
            _section(f"👩‍🏫  {who}{tag}",
                     grid_html(teacher_grid(m, solution, who, class_color), class_color))
    else:
        st.info("Click **Generate timetable** to view results.")

with tab_report:
    if res:
        m, solution, status, obj, errors, warnings = res
        c1, c2, c3 = st.columns(3)
        c1.metric("Status", status)
        c2.metric("Hard errors", len(errors))
        c3.metric("Warnings & notes", len(warnings))
        if errors:
            st.error("Hard-rule errors:\n\n" + "\n".join(f"- {e}" for e in errors))
        else:
            st.success("No hard-rule violations — class & teacher timetables tally.")

        load = defaultdict(int)
        for (c, d, p), (s, t) in solution.items():
            load[t] += 1
        for c in m.study_hour_classes:
            t = m.study_supervisor.get(c)
            if t:
                load[t] += m.study_days()
        ldf = (pd.DataFrame({"Teacher": list(load), "Periods/week": list(load.values())})
               .sort_values("Periods/week", ascending=False).reset_index(drop=True))
        st.subheader("Teacher weekly load")
        st.bar_chart(ldf.set_index("Teacher"))

        if warnings:
            with st.expander(f"Notes & leisure warnings ({len(warnings)})"):
                for w in warnings:
                    st.write("• " + w)

        dl1, dl2 = st.columns(2)
        with open(st.session_state.out, "rb") as fbin:
            dl1.download_button("⬇️  Download Excel (.xlsx)", fbin.read(),
                                file_name=f"{m.cfg.name}_Timetable_Final.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                width="stretch")
        if st.session_state.get("pdf"):
            with open(st.session_state.pdf, "rb") as fbin:
                dl2.download_button("⬇️  Download PDF (all classes + teachers)", fbin.read(),
                                    file_name=f"{m.cfg.name}_Timetable.pdf",
                                    mime="application/pdf", width="stretch")
    else:
        st.info("Click **Generate timetable** to view the report.")
